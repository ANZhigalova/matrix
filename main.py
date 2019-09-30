import numpy as np
import openpyxl as xl
from matplotlib import pyplot as pl
import os


def get_data():
    print("Введите размеры матрицы(только положительные числа): ", end=" ")
    lin = int(input())
    col = int(input())
    return lin, col


def get_matrix(lin, col):
    mtr = np.random.rand(lin, col)
    mtr *= 20
    mtr -= 10
    return mtr


def sum_m():
    if (m1.get_data == m2.get_data) and (n1.get_data == n2.get_data):
        return A.get_matrix + B.get_matrix
    else:
        print("невозможно сложить матрицы")


def mul_m():
    if n1.get_data == m2.get_data:
        return A.get_matrix@B.get_matrix
    else:
        print("невозможно перемножить матрицы")


def rec_m(mtr: np.array, ind):
    wb = xl.Workbook()
    if wb.sheetnames[0] == "Sheet":
        wb.remove(wb.sheetnames)
    sheet = wb.create_sheet(title='Матрица ', index=ind)
    for i in mtr:
        sheet.append(list(i))


def sum_col(sheet):
    y = []
    x = []
    if os.path.exists(sheet):
        wb = xl.Workbook(sheet, data_only=True)
        lines = sheet.max_row
        columns = sheet.max_column
        a = []
        for i in range(columns):
            x.append(i)
            s = 0
            for j in range(lines):
                s += a[i][j]
            y.append(s)
        return x, y
    else:
        print("Матрица не была построена")


def plot(x, y, e):
    f = pl.figure()
    f1 = f.add_subplot(2, 2, e)
    f1.plot(x, y, "go")
    pl.savefig("", fmt='png')
    pl.show()


m1, n1 = get_data()
m2, n2 = get_data()

A = get_matrix(m1, n1)
B = get_matrix(m2, n2)
S = sum_m(A, B)
M = mul_m(A, B)

rec_m(A, 1)
rec_m(B, 2)
rec_m(S, 3)
rec_m(M, 4)

a1, a2 = sum_col(["Матрица 1"])
b1, b2 = sum_col(["Матрица 2"])
s1, s2 = sum_col(["Матрица 3"])
m1, m2 = sum_col(["Матрица 4"])

plot(a1, a2, 1)
plot(b1, b2, 2)
plot(s1, s2, 3)
plot(m1, m2, 4)












